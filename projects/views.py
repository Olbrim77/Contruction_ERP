# projects/views.py

from decimal import Decimal
import math
import os
import re
from datetime import timedelta, datetime # <--- datetime hozzáadva

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders
from django.core.files.storage import default_storage
from django.db import transaction
# JAVÍTÁS: Itt vannak a hiányzó importok a fájl elején, nem a közepén!
from django.db.models import Sum, F, Q, Count
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import get_template
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse

# Modellek importálása (HR modellekkel bővítve)
from .models import (
    Project, Task, Tetelsor, Munkanem, Alvallalkozo, Expense, DailyLog,
    Supplier, Material, MasterItem, ItemComponent, CompanySettings,
    ProjectDocument, MaterialOrder, OrderItem, ProjectInventory, DailyMaterialUsage,
    GanttLink, UniclassNode, LaborComponent, MachineComponent, Operation, Machine,
    DailyLogImage, ProjectChapter,
    Employee, Attendance, PayrollItem # <-- Ezek kellenek a HR dashboardhoz
)

# Űrlapok importálása
from .forms import (
    ProjectForm, TetelsorQuantityForm, TetelsorEditForm, ExpenseForm,
    DailyLogForm, TetelsorCreateFromMasterForm, MasterItemForm, ItemComponentForm,
    ProjectDocumentForm, MaterialOrderForm, OrderItemFormSet, DailyMaterialUsageFormSet,
    TaskForm, LaborComponentForm, MachineComponentForm, MobilePhotoForm,
    MaterialInlineFormSet, LaborInlineFormSet, MachineInlineFormSet,
    DailyLogImageFormSet, ProjectDocumentFormSet, ProjectChapterForm
)

try:
    from xhtml2pdf import pisa
except ImportError:
    pisa = None

# Űrlapok importálása
from .forms import (
    ProjectForm, TetelsorQuantityForm, TetelsorEditForm, ExpenseForm,
    DailyLogForm, TetelsorCreateFromMasterForm, MasterItemForm, ItemComponentForm,
    ProjectDocumentForm, MaterialOrderForm, OrderItemFormSet, DailyMaterialUsageFormSet,
    TaskForm, LaborComponentForm, MachineComponentForm, MobilePhotoForm,
    MaterialInlineFormSet, LaborInlineFormSet, MachineInlineFormSet,
    DailyLogImageFormSet, ProjectDocumentFormSet,ProjectChapterForm
)

try:
    from xhtml2pdf import pisa
except ImportError:
    pisa = None

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


# --- SEGÉDFÜGGVÉNYEK ---------------------------------------------------------

def natural_sort_key(s):
    """
    Szöveges sorszámok helyes rendezése (pl. 1, 2, 10 a helyes, nem 1, 10, 2).
    """
    if not s:
        return []
    # A szöveget számokra és nem-számokra bontjuk
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', str(s))]

def calculate_work_end_date(start_date, workdays):
    """Munkanapok (H–P) alapján számolja ki a befejezési dátumot."""
    if not start_date or workdays <= 0:
        return None

    current_date = start_date
    days_to_add = int(workdays) - 1

    # Ha hétvégére esik a kezdés, toljuk hétfőre
    while current_date.weekday() >= 5:
        current_date += timedelta(days=1)

    added_days = 0
    while added_days < days_to_add:
        current_date += timedelta(days=1)
        if current_date.weekday() < 5:
            added_days += 1

    return current_date

def get_next_workday(date_obj):
    """ Ha hétvége, hétfőre tolja. Ha munkanap, marad. """
    while date_obj.weekday() >= 5:
        date_obj += timedelta(days=1)
    return date_obj


def get_cell_value(row, index, default=""):
    try:
        val = row[index].value
        return str(val).strip() if val is not None else default
    except Exception:
        return default


def get_cell_decimal(row, index):
    try:
        val = row[index].value
        if val is None:
            return Decimal(0)
        if isinstance(val, (int, float, Decimal)):
            return Decimal(val)
        val_str = (
            str(val)
            .replace("\xa0", "")
            .replace(" ", "")
            .replace("Ft", "")
            .replace(",", ".")
        )
        return Decimal(val_str) if val_str else Decimal(0)
    except Exception:
        return Decimal(0)


def get_next_version_id(base_id):
    """Új verziózott tetelszam (pl. BASE-E_001, BASE-E_002...)."""
    counter = 1
    while True:
        new_id = f"{base_id}-E_{counter:03d}"
        if not MasterItem.objects.filter(tetelszam=new_id).exists():
            return new_id
        counter += 1


def get_company_context():
    settings_obj = CompanySettings.objects.first()
    if settings_obj:
        logo_path = None
        if settings_obj.logo:
            try:
                # Ellenőrizzük, hogy létezik-e a fájl, nehogy IOError legyen
                logo_path = settings_obj.logo.path
                logo_path = logo_path.replace('\\', '/')
            except Exception:
                logo_path = None

        return {
            'company_name': settings_obj.name,
            'company_address': settings_obj.full_address(),
            'company_tax': settings_obj.tax_number,
            'company_phone': settings_obj.phone,
            'company_email': settings_obj.email,
            'company_logo': settings_obj.logo,
            'company_logo_path': logo_path,
            'signatories': settings_obj.signatories.all(),
            'default_city': settings_obj.head_city,
            'company_sites': settings_obj.sites.all(),
        }
    return {
        'company_name': 'Saját Kft.',
        'company_address': '-',
        'company_tax': '-',
        'company_phone': '-',
        'company_email': '-',
        'company_logo': None,
        'company_logo_path': None,
        'signatories': [],
        'default_city': 'Budapest',
        'company_sites': [],
    }


def link_callback(uri, rel):
    if uri.startswith('data:') or os.path.exists(uri):
        return uri
    result = finders.find(uri)
    if result:
        return result[0] if isinstance(result, (list, tuple)) else result
    return uri


# --- FŐ NÉZETEK ------------------------------------------------------------

@login_required
def project_list(request):
    projects = Project.objects.exclude(status='TORLES_KERELEM').order_by('start_date')
    q = request.GET.get('q')

    if q:
        projects = projects.filter(Q(name__icontains=q) | Q(location__icontains=q) | Q(client__icontains=q))

    project_groups = []
    if q:
        if projects.exists():
            project_groups.append({'label': f"Keresés: '{q}'", 'projects': projects})
    else:
        for sc, sl in Project.STATUS_CHOICES:
            projs = projects.filter(status=sc)
            if projs.exists():
                project_groups.append({'label': sl, 'projects': projs})

    all_budgets = Tetelsor.objects.filter(project__in=projects).aggregate(
        mat=Sum('anyag_osszesen'),
        ls=Sum('sajat_munkadij_osszesen'),
        la=Sum('alv_munkadij_osszesen'),
    )
    global_plan = (all_budgets['mat'] or 0) + (all_budgets['ls'] or 0) + (all_budgets['la'] or 0)

    global_spent = Expense.objects.filter(project__in=projects).aggregate(
        s=Sum('amount_netto')
    )['s'] or 0
    global_balance = global_plan - global_spent

    tasks = Task.objects.filter(status='FUGGO').order_by('due_date')

    if request.method == 'POST' and 'task_submit' in request.POST:
        task_form = TaskForm(request.POST)
        if task_form.is_valid():
            task_form.save()
            messages.success(request, "Feladat hozzáadva!")
            return redirect('project-list')
    else:
        task_form = TaskForm()

    status_counts = {
        'TERVEZES': projects.filter(
            status__in=['UJ_KERES', 'FELMERES', 'AJANLAT', 'ELOKESZITES']
        ).count(),
        'FOLYAMATBAN': projects.filter(status='KIVITELEZES').count(),
        'BEFEJEZETT': projects.filter(status='ATADAS').count(),
        'LEZART': projects.filter(status__in=['LEZART', 'ELUTASITVA']).count(),
    }

    context = {
        'project_groups': project_groups,
        'global_plan': global_plan,
        'global_spent': global_spent,
        'global_balance': global_balance,
        'status_counts': status_counts,
        'search_query': q or "",
        'tasks': tasks,
        'task_form': task_form,
    }
    return render(request, 'projects/project_list.html', context)


@login_required
def task_complete(request, pk):
    task = get_object_or_404(Task, id=pk)
    task.status = 'KESZ'
    task.save()
    messages.success(request, "Feladat elvégezve!")
    return redirect('project-list')


# --- PROJEKT ADATLAP ----------------------------------------------------------

@login_required
def project_detail(request, pk):
    project = get_object_or_404(Project, id=pk)

    # --- 1. ADATOK LEKÉRÉSE ---
    tasks = project.tetelsorok.all().order_by('sorszam')
    expenses = project.expenses.all().order_by('-date')
    documents = project.documents.all().order_by('-uploaded_at')
    orders = project.material_orders.all().order_by('-date')

    # Raktárkészlet (hibatűréssel, ha még nincs inventory relation)
    try:
        inventory = project.inventory.all().order_by('name')
    except AttributeError:
        inventory = []

    # --- 2. SZŰRÉS ---
    if request.GET.get('expense_category_filter'):
        expenses = expenses.filter(category=request.GET.get('expense_category_filter'))

    if request.GET.get('munkanem_filter'):
        tasks = tasks.filter(munkanem__nev=request.GET.get('munkanem_filter'))

    # Listák a legördülőkhöz
    munkanem_list = Munkanem.objects.all().order_by('nev')
    expense_category_list = [('ANYAG', 'Anyag'), ('MUNKADIJ', 'Munkadíj'), ('EGYEB', 'Egyéb')]

    # --- 3. PÉNZÜGYI ÖSSZESÍTÉS (AGGREGÁCIÓ) ---
    summ = tasks.aggregate(
        mat=Sum('anyag_osszesen'),
        ls=Sum('sajat_munkadij_osszesen'),
        la=Sum('alv_munkadij_osszesen'),
        hours=Sum(F('mennyiseg') * F('normaido')),
    )

    # Tervezett költségek (ha nincs adat, akkor 0)
    plan_mat = summ['mat'] or 0
    plan_ls = summ['ls'] or 0
    plan_la = summ['la'] or 0
    plan_total = plan_mat + plan_ls + plan_la

    # Tényleges kiadások
    spent = expenses.aggregate(s=Sum('amount_netto'))['s'] or 0

    # Egyenleg (Maradék)
    balance = plan_total - spent

    # --- 4. IDŐTARTAM SZÁMÍTÁS (ITT VOLT A HIBA) ---
    hours = summ['hours'] or 0

    # Biztonságos osztás: ha a hours_per_day nincs megadva, alapértelmezett 8
    hpd = project.hours_per_day if project.hours_per_day else 8

    # Itt használjuk a math modult tisztán
    if hpd > 0:
        days = math.ceil(hours / hpd)
    else:
        days = 0

    # Befejezés dátuma (Munkanapok alapján)
    end_date = calculate_work_end_date(project.start_date, days)

    # ÁFA és Bruttó számítás
    vat_rate = project.vat_rate or 27
    total_vat = plan_total * (vat_rate / 100)
    total_project_brutto = plan_total * (1 + vat_rate / 100)

    # --- 5. CONTEXT ÖSSZEÁLLÍTÁSA ---
    context = {
        'project': project,
        'tasks': tasks,
        'expenses': expenses,
        'documents': documents,
        'orders': orders,
        'inventory': inventory,

        # Pénzügy
        'plan_anyag': plan_mat,
        'plan_dij': plan_ls + plan_la,
        'plan_total': plan_total,
        'actual_total': spent,
        'balance': balance,
        'vat_rate': vat_rate,
        'total_vat': total_vat,
        'total_project_brutto': total_project_brutto,

        # Idő
        'total_effort_hours': hours,
        'total_workdays': days,
        'calculated_end_date': end_date,

        # Szűrők és Státuszok
        'munkanem_list': munkanem_list,
        'expense_category_list': expense_category_list,
        'current_munkanem': request.GET.get('munkanem_filter'),
        'current_expense_category': request.GET.get('expense_category_filter'),
        'all_statuses': Project.STATUS_CHOICES,
    }

    return render(request, 'projects/project_detail.html', context)

@login_required
def project_status_update(request, pk, status_code):
    project = get_object_or_404(Project, id=pk)
    valid_codes = [c[0] for c in Project.STATUS_CHOICES]

    if status_code in valid_codes:
        project.status = status_code
        project.save()
        messages.success(request, f"Státusz frissítve: {project.get_status_display()}")
    else:
        messages.error(request, "Érvénytelen státusz!")

    return redirect('project-detail', pk=project.id)


# --- PROJEKT CRUD -------------------------------------------------------------

@login_required
def project_create(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            new_project = form.save()
            return redirect('project-detail', pk=new_project.id)
    else:
        form = ProjectForm()
    return render(request, 'projects/project_form.html', {'form': form})


@login_required
def project_update(request, pk):
    project = get_object_or_404(Project, id=pk)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            # ha változott az óradíj, újraszámoljuk a tételsorokat
            if 'hourly_rate' in form.changed_data:
                for t in project.tetelsorok.all():
                    t.save()
            return redirect('project-detail', pk=project.id)
    else:
        form = ProjectForm(instance=project)
    return render(request, 'projects/project_form.html', {'form': form, 'project': project})


@login_required
def project_request_deletion(request, pk):
    project = get_object_or_404(Project, id=pk)
    if request.method == 'POST':
        project.status = 'TORLES_KERELEM'
        project.save()
        return redirect('project-list')
    context = {'project': project}
    return render(request, 'projects/project_confirm_delete.html', context)


# --- TÉTELSOR MÓDOSÍTÁS -------------------------------------------------------

@login_required
def tetelsor_update_quantity(request, pk):
    tetelsor = get_object_or_404(Tetelsor, pk=pk)

    if request.method == "POST":
        form = TetelsorQuantityForm(request.POST, instance=tetelsor)
        if form.is_valid():
            form.save()
            return redirect('project-detail', pk=tetelsor.project_id)
    else:
        form = TetelsorQuantityForm(instance=tetelsor)

    return render(
        request,
        'projects/tetelsor_form.html',
        {
            'form': form,
            'tetelsor': tetelsor,
        }
    )


@login_required
def tetelsor_update(request, pk):
    tetelsor = get_object_or_404(Tetelsor, pk=pk)
    # JAVÍTÁS: Lekérjük a projektet, hogy át tudjuk adni a sablonnak
    project = tetelsor.project

    if request.method == "POST":
        form = TetelsorEditForm(request.POST, instance=tetelsor)
        if form.is_valid():
            form.save()
            messages.success(request, "Tétel módosítva.")
            # JAVÍTÁS: Visszairányítás a Költségvetés (#budget) fülre
            url = reverse('project-detail', kwargs={'pk': project.id}) + '#budget'
            return redirect(url)
    else:
        form = TetelsorEditForm(instance=tetelsor)

    return render(
        request,
        'projects/tetelsor_edit_form.html',
        {
            'form': form,
            'tetelsor': tetelsor,
            'project': project,  # <--- EZ HIÁNYZOTT, EMIATT VOLT A HIBA!
            'submit_label': 'Mentés'
        }
    )


@login_required
def tetelsor_create_from_master_step2(request, project_id, master_id):
    project = get_object_or_404(Project, id=project_id)
    master = get_object_or_404(MasterItem, id=master_id)

    # Sorszám logika
    last_item = Tetelsor.objects.filter(project=project).last()
    new_sorszam = 1
    if last_item and last_item.sorszam.isdigit():
        new_sorszam = int(last_item.sorszam) + 1

    t = Tetelsor(
        project=project, master_item=master, sorszam=str(new_sorszam),
        mennyiseg=0, anyag_egysegar=master.fix_anyag_ar, leiras=master.leiras,
        egyseg=master.egyseg, normaido=master.normaido, munkanem=master.munkanem,
        engy_kod=master.engy_kod, k_jelzo=master.k_jelzo, cpr_kod=master.cpr_kod
    )

    if request.method == 'POST':
        form = TetelsorEditForm(request.POST, instance=t)
        if form.is_valid():
            tetel = form.save(commit=False)
            tetel.project = project
            tetel.master_item = master
            tetel.save()
            messages.success(request, "Tétel beillesztve!")
            # JAVÍTÁS: Visszairányítás a #budget fülre
            url = reverse('project-detail', kwargs={'pk': project.id}) + '#budget'
            return redirect(url)
    else:
        form = TetelsorEditForm(instance=t)

    return render(request, 'projects/tetelsor_edit_form.html', {
        'form': form,
        'project': project,  # Itt át volt adva, ez jó volt
        'tetelsor': t,
        'submit_label': 'Beillesztés'
    })


@login_required
def tetelsor_delete(request, pk):
    tetelsor = get_object_or_404(Tetelsor, pk=pk)
    project_id = tetelsor.project_id

    if request.method == "POST":
        tetelsor.delete()
        messages.success(request, "Tétel törölve.")
        return redirect('project-detail', pk=project_id)

    return render(
        request,
        'projects/tetelsor_confirm_delete.html',
        {'tetelsor': tetelsor}
    )


# --- TÉTELSOR IMPORT ---------------------------------------------------------

@login_required
def import_tasks(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST' and 'temp_file_path' in request.POST:
        path = os.path.join(settings.MEDIA_ROOT, request.POST.get('temp_file_path'))
        try:
            # JAVÍTÁS: Explicit lezárás és try-except
            wb = openpyxl.load_workbook(path, data_only=True)
            cnt = 0
            for sheet_name in request.POST.getlist('sheets'):
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2):
                    tid = get_cell_value(row, 1)
                    if not tid: continue

                    desc = get_cell_value(row, 2)
                    qty = get_cell_decimal(row, 3)
                    unit = get_cell_value(row, 4)
                    price = get_cell_decimal(row, 5)

                    # JAVÍTÁS: Üres mezők ellenőrzése
                    mn_nev = get_cell_value(row, 12).strip()
                    mn = None
                    if mn_nev:
                        mn, _ = Munkanem.objects.get_or_create(nev=mn_nev)

                    alv_nev = get_cell_value(row, 14).strip()
                    alv = None
                    if alv_nev:
                        alv, _ = Alvallalkozo.objects.get_or_create(nev=alv_nev)

                    mi = MasterItem.objects.filter(tetelszam=tid).first()
                    if not mi:
                        mi = MasterItem.objects.create(
                            tetelszam=tid, leiras=desc, egyseg=unit,
                            normaido=get_cell_decimal(row, 13), fix_anyag_ar=price, munkanem=mn
                        )
                    elif mi.leiras.strip() != desc.strip():
                        mi = MasterItem.objects.create(
                            tetelszam=get_next_version_id(tid), leiras=desc, egyseg=unit,
                            normaido=get_cell_decimal(row, 13), fix_anyag_ar=price, munkanem=mn
                        )

                    Tetelsor.objects.update_or_create(
                        project=project, master_item=mi,
                        defaults={
                            'sorszam': str(cnt), 'mennyiseg': qty, 'anyag_egysegar': price,
                            'leiras': desc, 'egyseg': unit, 'normaido': get_cell_decimal(row, 13),
                            'munkanem': mn, 'alvallalkozo': alv
                        }
                    )
                    cnt += 1

            wb.close()  # FONTOS: Bezárás
            messages.success(request, f"Sikeres import! ({cnt} sor)")

        except Exception as e:
            messages.error(request, f"Hiba: {e}")
        finally:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass
        return redirect('project-detail', pk=project.id)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        f = request.FILES['excel_file']
        path = default_storage.save(f"temp/{f.name}", f)

        # Csak a lapnevek miatt nyitjuk meg, majd bezárjuk
        wb = openpyxl.load_workbook(os.path.join(settings.MEDIA_ROOT, path), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        return render(request, 'projects/import_step2_select.html', {
            'project': project, 'sheet_names': sheet_names, 'temp_file_path': path
        })

    return render(request, 'projects/import_step1_upload.html', {'project': project})


@login_required
def tetelsor_create_from_master_step1(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    items = MasterItem.objects.all().order_by('tetelszam')
    q = request.GET.get('q')
    mn = request.GET.get('munkanem')
    if q:
        items = items.filter(Q(leiras__icontains=q) | Q(tetelszam__icontains=q))
    if mn:
        items = items.filter(munkanem__id=mn)
    return render(request, 'projects/tetelsor_create_from_master.html',
                  {'project': project, 'items': items, 'munkanemek': Munkanem.objects.all().order_by('nev'),
                   'search_query': q or "", 'current_munkanem': int(mn) if mn else ""})


@login_required
def tetelsor_finalize_new(request, project_id, master_id):
    return tetelsor_create_from_master_step2(request, project_id, master_id)


@login_required
def tetelsor_create_from_master_step2(request, project_id, master_id):
    project = get_object_or_404(Project, id=project_id)
    master = get_object_or_404(MasterItem, id=master_id)

    # ÚJ sorszám keresése
    last_item = Tetelsor.objects.filter(project=project).last()
    new_sorszam = 1
    if last_item and last_item.sorszam.isdigit():
        new_sorszam = int(last_item.sorszam) + 1

    t = Tetelsor(
        project=project, master_item=master, sorszam=str(new_sorszam),
        mennyiseg=0, anyag_egysegar=master.fix_anyag_ar, leiras=master.leiras, egyseg=master.egyseg,
        normaido=master.normaido, munkanem=master.munkanem,
        engy_kod=master.engy_kod, k_jelzo=master.k_jelzo, cpr_kod=master.cpr_kod
    )

    if request.method == 'POST':
        form = TetelsorEditForm(request.POST, instance=t)
        if form.is_valid():
            tetel = form.save(commit=False)
            tetel.project = project
            tetel.master_item = master
            tetel.save()
            messages.success(request, "Tétel beillesztve!")
            # JAVÍTÁS: Visszairányítás a Költségvetés fülre (#budget)
            return redirect(reverse('project-detail', kwargs={'pk': project.id}) + '#budget')
    else:
        form = TetelsorEditForm(instance=t)

    return render(request, 'projects/tetelsor_edit_form.html', {
        'form': form,
        'project': project,
        'tetelsor': t,
        'submit_label': 'Beillesztés'
    })


@login_required
def sync_tetelsor_to_master(request, pk):
    tetelsor = get_object_or_404(Tetelsor, id=pk)
    if not tetelsor.master_item:
        messages.error(request, "Ez a tétel nem csatolt a törzshöz!")
        return redirect('project-detail', pk=tetelsor.project.id)
    master = tetelsor.master_item
    master.leiras = tetelsor.leiras;
    master.normaido = tetelsor.normaido;
    master.fix_anyag_ar = tetelsor.anyag_egysegar;
    master.egyseg = tetelsor.egyseg
    master.save()
    messages.success(request, "Törzselem frissítve!")
    return redirect('project-detail', pk=tetelsor.project.id)


# --- DOKUMENTUMOK (FORMSET) ---

@login_required
def document_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        formset = ProjectDocumentFormSet(request.POST, request.FILES, queryset=ProjectDocument.objects.none())
        if formset.is_valid():
            # JAVÍTÁS: Tranzakcióban, hogy atomi legyen
            try:
                with transaction.atomic():
                    documents = formset.save(commit=False)
                    for doc in documents:
                        doc.project = project
                        doc.save()
                messages.success(request, f"{len(documents)} dokumentum feltöltve.")
                return redirect('project-detail', pk=project.id)
            except Exception as e:
                messages.error(request, f"Hiba történt a mentéskor: {e}")
    else:
        formset = ProjectDocumentFormSet(queryset=ProjectDocument.objects.none())
    return render(request, 'projects/document_form.html', {'formset': formset, 'project': project})


@login_required
def document_delete(request, pk):
    doc = get_object_or_404(ProjectDocument, pk=pk);
    pid = doc.project.id
    if request.method == 'POST': doc.delete(); messages.success(request, "Törölve."); return redirect('project-detail',
                                                                                                      pk=pid)
    return render(request, 'projects/document_confirm_delete.html', {'document': doc})


# --- NAPI NAPLÓ (JAVÍTOTT KÉSZLETKEZELÉSSEL) ---

@login_required
def daily_log_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        form = DailyLogForm(request.POST)
        mat_formset = DailyMaterialUsageFormSet(request.POST, prefix='materials')
        img_formset = DailyLogImageFormSet(request.POST, request.FILES, prefix='images')

        if form.is_valid() and mat_formset.is_valid() and img_formset.is_valid():
            try:
                with transaction.atomic():
                    log = form.save(commit=False)
                    log.project = project
                    log.save()

                    # Anyagok (JAVÍTOTT: F() expression a Race Condition ellen)
                    usages = mat_formset.save(commit=False)
                    for usage in usages:
                        usage.log = log
                        usage.save()
                        if usage.inventory_item:
                            # Atomikus frissítés
                            ProjectInventory.objects.filter(pk=usage.inventory_item.pk).update(
                                quantity=F('quantity') - usage.quantity
                            )

                    # Képek
                    images = img_formset.save(commit=False)
                    for img in images:
                        img.log = log
                        img.save()

                messages.success(request, "Napló rögzítve.")
                return redirect('project-detail', pk=project.id)
            except Exception as e:
                messages.error(request, f"Hiba a mentés során: {e}")
    else:
        try:
            existing = DailyLog.objects.get(project=project, date=timezone.localdate())
            return redirect('daily-log-update', pk=existing.id)
        except DailyLog.DoesNotExist:
            form = DailyLogForm(initial={'date': timezone.localdate()})
            mat_formset = DailyMaterialUsageFormSet(prefix='materials', form_kwargs={'project': project})
            img_formset = DailyLogImageFormSet(prefix='images')

    return render(request, 'projects/daily_log_form.html', {
        'form': form, 'mat_formset': mat_formset, 'img_formset': img_formset, 'project': project
    })


@login_required
def daily_log_update(request, pk):
    log = get_object_or_404(DailyLog, pk=pk)
    project = log.project

    if request.method == 'POST':
        form = DailyLogForm(request.POST, instance=log)
        mat_formset = DailyMaterialUsageFormSet(request.POST, instance=log, prefix='materials',
                                                form_kwargs={'project': project})
        img_formset = DailyLogImageFormSet(request.POST, request.FILES, instance=log, prefix='images')

        if form.is_valid() and mat_formset.is_valid() and img_formset.is_valid():
            try:
                with transaction.atomic():
                    log = form.save()

                    # Anyagok (JAVÍTOTT: Különbség kezelése és Race Condition ellen)
                    instances = mat_formset.save(commit=False)
                    for usage in instances:
                        if usage.pk:
                            # Meglévő rekord: lekérjük a régit a DB-ből
                            try:
                                old_usage = DailyMaterialUsage.objects.get(pk=usage.pk)
                                diff = usage.quantity - old_usage.quantity
                                if usage.inventory_item:
                                    ProjectInventory.objects.filter(pk=usage.inventory_item.pk).update(
                                        quantity=F('quantity') - diff
                                    )
                            except DailyMaterialUsage.DoesNotExist:
                                pass
                        else:
                            # Új rekord
                            if usage.inventory_item:
                                ProjectInventory.objects.filter(pk=usage.inventory_item.pk).update(
                                    quantity=F('quantity') - usage.quantity
                                )

                        usage.log = log
                        usage.save()

                    # Törölt sorok kezelése (visszavételezés)
                    for obj in mat_formset.deleted_objects:
                        if obj.inventory_item:
                            ProjectInventory.objects.filter(pk=obj.inventory_item.pk).update(
                                quantity=F('quantity') + obj.quantity
                            )
                        obj.delete()

                    # Képek
                    img_formset.save()

                messages.success(request, "Napló módosítva.")
                return redirect('project-detail', pk=project.id)
            except Exception as e:
                messages.error(request, f"Hiba a mentés során: {e}")

    else:
        form = DailyLogForm(instance=log)
        mat_formset = DailyMaterialUsageFormSet(instance=log, prefix='materials', form_kwargs={'project': project})
        img_formset = DailyLogImageFormSet(instance=log, prefix='images')

    return render(request, 'projects/daily_log_form.html', {
        'form': form, 'mat_formset': mat_formset, 'img_formset': img_formset, 'log_entry': log, 'project': project
    })


@login_required
def daily_log_delete(request, pk):
    l = get_object_or_404(DailyLog, id=pk);
    pid = l.project.id
    if request.method == 'POST': l.delete(); messages.success(request, "Napló törölve."); return redirect(
        'project-detail', pk=pid)
    return render(request, 'projects/daily_log_confirm_delete.html', {'log_entry': l})


@login_required
def daily_log_image_delete(request, pk):
    img = get_object_or_404(DailyLogImage, pk=pk);
    log_id = img.log.id
    if request.method == 'POST': img.delete(); messages.success(request, "Fotó törölve.")
    return redirect('daily-log-update', pk=log_id)


@login_required
def daily_log_detail(request, pk):
    log = get_object_or_404(DailyLog, pk=pk)
    return render(request, 'projects/daily_log_detail.html', {
        'log': log, 'project': log.project, 'images': log.images.all(), 'materials': log.material_usages.all()
    })


# --- MASTER ITEM CRUD (Receptúrával) ---

@login_required
def master_item_list(request):
    items = MasterItem.objects.all().order_by('tetelszam');
    q = request.GET.get('q')
    if q: items = items.filter(Q(leiras__icontains=q) | Q(tetelszam__icontains=q))
    return render(request, 'projects/master_item_list.html', {'items': items, 'munkanemek': Munkanem.objects.all()})


@login_required
def master_item_create(request):
    if request.method == 'POST':
        form = MasterItemForm(request.POST)
        mat_formset = MaterialInlineFormSet(request.POST, prefix='materials')
        lab_formset = LaborInlineFormSet(request.POST, prefix='labor')
        mach_formset = MachineInlineFormSet(request.POST, prefix='machines')

        if form.is_valid() and mat_formset.is_valid() and lab_formset.is_valid() and mach_formset.is_valid():
            try:
                with transaction.atomic():
                    master_item = form.save()
                    mat_formset.instance = master_item;
                    mat_formset.save()
                    lab_formset.instance = master_item;
                    lab_formset.save()
                    mach_formset.instance = master_item;
                    mach_formset.save()
                    master_item.calculate_totals()
                messages.success(request, "Új tétel és receptúra létrehozva!")
                return redirect('master-item-list')
            except Exception as e:
                messages.error(request, f"Hiba: {e}")
    else:
        form = MasterItemForm()
        mat_formset = MaterialInlineFormSet(prefix='materials')
        lab_formset = LaborInlineFormSet(prefix='labor')
        mach_formset = MachineInlineFormSet(prefix='machines')

    return render(request, 'projects/master_item_form.html',
                  {'form': form, 'mat_formset': mat_formset, 'lab_formset': lab_formset, 'mach_formset': mach_formset,
                   'title': 'Új Törzs Tétel'})


@login_required
def master_item_update(request, pk):
    item = get_object_or_404(MasterItem, pk=pk)
    if request.method == 'POST':
        form = MasterItemForm(request.POST, instance=item)
        mat_formset = MaterialInlineFormSet(request.POST, instance=item, prefix='materials')
        lab_formset = LaborInlineFormSet(request.POST, instance=item, prefix='labor')
        mach_formset = MachineInlineFormSet(request.POST, instance=item, prefix='machines')

        if form.is_valid() and mat_formset.is_valid() and lab_formset.is_valid() and mach_formset.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    mat_formset.save();
                    lab_formset.save();
                    mach_formset.save()
                    item.calculate_totals()
                messages.success(request, "Tétel frissítve.")
                return redirect('master-item-list')
            except Exception as e:
                messages.error(request, f"Hiba: {e}")
    else:
        form = MasterItemForm(instance=item)
        mat_formset = MaterialInlineFormSet(instance=item, prefix='materials')
        lab_formset = LaborInlineFormSet(instance=item, prefix='labor')
        mach_formset = MachineInlineFormSet(instance=item, prefix='machines')

    return render(request, 'projects/master_item_form.html',
                  {'form': form, 'item': item, 'mat_formset': mat_formset, 'lab_formset': lab_formset,
                   'mach_formset': mach_formset, 'title': 'Tétel Szerkesztése'})


@login_required
def master_item_delete(request, pk):
    i = get_object_or_404(MasterItem, pk=pk)
    if request.method == 'POST': i.delete(); messages.success(request, "Törölve."); return redirect('master-item-list')
    return render(request, 'projects/master_item_confirm_delete.html', {'item': i})


@login_required
def master_item_components(request, pk): return redirect('master-item-update', pk=pk)


@login_required
def master_item_component_delete(request, pk): return redirect('master-item-list')  # Nem kell, formset kezeli


@login_required
def import_master_items(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file'];
        wb = openpyxl.load_workbook(excel_file, data_only=True);
        ws = wb.active
        created = 0
        for row in ws.iter_rows(min_row=2):
            code = get_cell_value(row, 0);
            desc = get_cell_value(row, 1)
            if code and desc:
                MasterItem.objects.update_or_create(tetelszam=code,
                                                    defaults={'leiras': desc, 'egyseg': get_cell_value(row, 2),
                                                              'normaido': get_cell_decimal(row, 3),
                                                              'fix_anyag_ar': get_cell_decimal(row, 4)})
                created += 1
        messages.success(request, f"{created} tétel importálva.");
        return redirect('master-item-list')
    return render(request, 'projects/master_item_import.html')


# --- KÖLTSÉG ---

@login_required
def expense_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            exp = form.save(commit=False);
            exp.project = project;
            exp.save()
            messages.success(request, "Kiadás rögzítve.")
            return redirect('project-detail', pk=project.id)
    else:
        form = ExpenseForm()
    return render(request, 'projects/expense_form.html', {'form': form, 'project': project})


@login_required
def expense_update(request, pk):
    e = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES, instance=e)
        if form.is_valid(): form.save(); messages.success(request, "Kiadás módosítva."); return redirect(
            'project-detail', pk=e.project.id)
    else:
        form = ExpenseForm(instance=e)
    return render(request, 'projects/expense_form.html', {'form': form, 'expense': e, 'project': e.project})


@login_required
def expense_delete(request, pk):
    e = get_object_or_404(Expense, pk=pk);
    pid = e.project.id
    if request.method == 'POST': e.delete(); messages.success(request, "Kiadás törölve."); return redirect(
        'project-detail', pk=pid)
    return render(request, 'projects/expense_confirm_delete.html', {'object': e})


# --- RENDELÉS ---

@login_required
def material_order_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        form = MaterialOrderForm(request.POST);
        formset = OrderItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            order = form.save(commit=False);
            order.project = project;
            order.save()
            items = formset.save(commit=False)
            for it in items: it.order = order; it.save()
            for obj in formset.deleted_objects: obj.delete()
            messages.success(request, "Rendelés létrehozva.")
            return redirect('project-detail', pk=project.id)
    else:
        form = MaterialOrderForm(); formset = OrderItemFormSet()
    return render(request, 'projects/material_order_form.html', {'form': form, 'formset': formset, 'project': project})


@login_required
def material_order_update(request, pk):
    order = get_object_or_404(MaterialOrder, pk=pk)
    if request.method == 'POST':
        form = MaterialOrderForm(request.POST, instance=order);
        formset = OrderItemFormSet(request.POST, instance=order)
        if form.is_valid() and formset.is_valid(): form.save(); formset.save(); messages.success(request,
                                                                                                 "Módosítva."); return redirect(
            'project-detail', pk=order.project.id)
    else:
        form = MaterialOrderForm(instance=order); formset = OrderItemFormSet(instance=order)
    return render(request, 'projects/material_order_form.html',
                  {'form': form, 'formset': formset, 'project': order.project, 'order': order})


@login_required
def material_order_delete(request, pk):
    order = get_object_or_404(MaterialOrder, pk=pk);
    pid = order.project.id
    if request.method == 'POST': order.delete(); return redirect('project-detail', pk=pid)
    return render(request, 'projects/document_confirm_delete.html', {'document': order})


@login_required
def material_order_create_from_budget(request, project_id):
    project = get_object_or_404(Project, id=project_id);
    tasks = project.tetelsorok.filter(anyag_egysegar__gt=0).order_by('sorszam')
    if request.method == 'POST':
        ids = request.POST.getlist('selected_items')
        if ids:
            order = MaterialOrder.objects.create(project=project, status='TERVEZET', notes="Költségvetésből")
            for tid in ids: t = Tetelsor.objects.get(id=tid); OrderItem.objects.create(order=order, name=t.leiras,
                                                                                       quantity=t.mennyiseg,
                                                                                       unit=t.egyseg,
                                                                                       price=t.anyag_egysegar)
            return redirect('material-order-update', pk=order.id)
    return render(request, 'projects/material_order_from_budget.html', {'project': project, 'tasks': tasks})


@login_required
def material_order_print(request, pk):
    order = get_object_or_404(MaterialOrder, id=pk);
    total = sum(item.total_price for item in order.items.all())
    return render(request, 'projects/material_order_print.html',
                  {'order': order, 'items': order.items.all(), 'total_value': total, 'today': timezone.now(),
                   **get_company_context()})


@login_required
def material_order_pdf(request, pk): return material_order_print(request, pk)


@login_required
def material_order_finalize(request, pk):
    order = get_object_or_404(MaterialOrder, id=pk)
    if request.method == 'POST':
        total = sum(i.total_price for i in order.items.all())
        if total > 0: Expense.objects.create(project=order.project, name=f"Rendelés #{order.id}", date=timezone.now(),
                                             category='ANYAG', amount_netto=total)
        for item in order.items.all():
            inv, _ = ProjectInventory.objects.get_or_create(project=order.project, name=item.name,
                                                            defaults={'unit': item.unit});
            inv.quantity += item.quantity;
            inv.save()
        order.status = 'TELJESITVE';
        order.save();
        messages.success(request, "Könyvelve!")
        return redirect('project-detail', pk=order.project.id)
    return render(request, 'projects/material_order_finalize.html', {'order': order})


# === GANTT DIAGRAM ===

def gantt_view(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    alvallalkozok = Alvallalkozo.objects.all().values('id', 'nev')
    return render(request, 'projects/gantt_view.html', {'project': project, 'alvallalkozok': list(alvallalkozok)})


def gantt_data(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    # 1. Lekérjük az összes tételt (még rendezetlenül)
    tasks_qs = project.tetelsorok.all()

    # 2. PYTHON RENDEZÉS (EZ A KULCS!)
    # Itt használjuk a fenti segédfüggvényt, hogy a '11' tényleg a '2' után jöjjön.
    tasks = sorted(tasks_qs, key=lambda t: natural_sort_key(t.sorszam or ""))

    data = []

    # 3. Lánc kezdőpontja (Mutató)
    # Alapból a projekt kezdete, vagy ha nincs, a mai nap
    current_pointer = project.start_date or timezone.now().date()

    for t in tasks:
        # --- A) IDŐTARTAM SZÁMÍTÁS ---
        calc_duration = 1
        if t.normaido and t.mennyiseg:
            try:
                hpd = float(project.hours_per_day or 8)
                if hpd > 0:
                    calc_duration = math.ceil((float(t.mennyiseg) * float(t.normaido)) / hpd)
            except Exception:
                pass

        duration = calc_duration

        # Ha van mentett kézi hossz (és > 0!), az felülírja a számítottat
        if t.gantt_start_date:
            if t.gantt_duration and t.gantt_duration > 0:
                duration = t.gantt_duration
            elif calc_duration > 0:
                duration = calc_duration

        if duration < 1: duration = 1

        # --- B) KEZDÉS DÁTUMA (VÍZESÉS) ---
        # Ha a felhasználó kézzel elhúzta, az felülírja a láncot
        if t.gantt_start_date:
            start_date_obj = t.gantt_start_date
        else:
            # Ha nem, akkor az előző feladat után jön
            start_date_obj = current_pointer

        # Ha a számított kezdés hétvégére esne, toljuk hétfőre
        while start_date_obj.weekday() >= 5:
            start_date_obj += timedelta(days=1)

        # --- C) BEFEJEZÉS ÉS ÚJ MUTATÓ ---
        finish_date_obj = calculate_work_end_date(start_date_obj, duration)

        # A mutatót a befejezés utáni napra állítjuk
        if finish_date_obj:
            next_start = finish_date_obj + timedelta(days=1)
            # A következő feladat startja se legyen hétvége
            while next_start.weekday() >= 5:
                next_start += timedelta(days=1)
            current_pointer = next_start
        else:
            current_pointer = start_date_obj + timedelta(days=1)

        # Adatok formázása
        start_date_str = start_date_obj.strftime("%Y-%m-%d")
        finish_date_str = finish_date_obj.strftime("%Y-%m-%d") if finish_date_obj else ""

        txt = t.leiras or (t.master_item.leiras if t.master_item else "-")
        tszam = (t.master_item.tetelszam if t.master_item else (t.sorszam if t.sorszam else ""))
        mn = t.munkanem.nev if t.munkanem else "-"

        data.append({
            'id': t.id,
            'tetelszam': tszam,
            'text': txt,
            'start_date': start_date_str,
            'finish_date': finish_date_str,
            'duration': duration,
            'progress': float(t.progress_percentage) / 100 if t.progress_percentage else 0,
            'open': True,
            'mennyiseg': f"{t.mennyiseg} {t.egyseg}",
            'munkanem': mn,
            'felelos': t.felelos or "",
            'owner_id': t.alvallalkozo.id if t.alvallalkozo else "",
        })

    link_data = []
    try:
        links = GanttLink.objects.filter(source__project=project)
        link_data = [{'id': l.id, 'source': l.source.id, 'target': l.target.id, 'type': l.type} for l in links]
    except Exception:
        pass

    return JsonResponse({"data": data, "links": link_data})


def global_gantt_data(request):
    projects = Project.objects.exclude(status='TORLES_KERELEM').order_by('start_date')
    data = []

    # Minden projektnek saját mutatója van a láncoláshoz
    project_pointers = {}

    for p in projects:
        start = p.start_date or timezone.now().date()
        project_pointers[p.id] = start

        data.append({
            'id': f"prj_{p.id}",
            'tetelszam': "PROJEKT",
            'text': p.name,
            'start_date': start.strftime("%Y-%m-%d"),
            'type': 'project',
            'open': False,
            'readonly': True,
        })

    # Összes feladat lekérése
    all_tasks = Tetelsor.objects.filter(project__in=projects)

    # RENDEZÉS: Először Projekt szerint, azon belül pedig Természetes Sorszám szerint
    tasks = sorted(all_tasks, key=lambda t: (t.project.id, natural_sort_key(t.sorszam or "")))

    for t in tasks:
        # Az adott projekt mutatójának lekérése
        current_pointer = project_pointers.get(t.project.id, timezone.now().date())

        calc_duration = 1
        if t.normaido and t.mennyiseg:
            try:
                hpd = float(t.project.hours_per_day or 8)
                if hpd > 0: calc_duration = math.ceil((float(t.mennyiseg) * float(t.normaido)) / hpd)
            except:
                pass

        duration = calc_duration
        if t.gantt_start_date:
            if t.gantt_duration and t.gantt_duration > 0:
                duration = t.gantt_duration
            elif calc_duration > 1:
                duration = calc_duration

        if duration < 1: duration = 1

        # Kezdés (Vízesés)
        if t.gantt_start_date:
            start_date_obj = t.gantt_start_date
        else:
            start_date_obj = current_pointer

        # Hétvége korrekció
        while start_date_obj.weekday() >= 5:
            start_date_obj += timedelta(days=1)

        finish_date_obj = calculate_work_end_date(start_date_obj, duration)

        # Mutató frissítése az adott projekthez
        if finish_date_obj:
            next_start = finish_date_obj + timedelta(days=1)
            while next_start.weekday() >= 5:
                next_start += timedelta(days=1)
            project_pointers[t.project.id] = next_start

        start_date_str = start_date_obj.strftime("%Y-%m-%d")
        txt = t.leiras or (t.master_item.leiras if t.master_item else "-")
        tszam = t.master_item.tetelszam if t.master_item else ""

        data.append({
            'id': t.id,
            'tetelszam': tszam,
            'text': txt,
            'start_date': start_date_str,
            'duration': duration,
            'progress': float(t.progress_percentage) / 100 if t.progress_percentage else 0,
            'parent': f"prj_{t.project.id}",
            'owner_id': t.alvallalkozo.id if t.alvallalkozo else "",
            'felelos': t.felelos or "",
        })

    links = GanttLink.objects.filter(source__project__in=projects)
    link_data = [{'id': l.id, 'source': l.source.id, 'target': l.target.id, 'type': l.type} for l in links]

    return JsonResponse({"data": data, "links": link_data})
@csrf_exempt
def gantt_update(request, project_id):
    if request.method == 'POST':
        try:
            mode = request.POST.get("gantt_mode"); op = request.POST.get("!nativeeditor_status"); sid = request.POST.get("id")
            is_link = ('source' in request.POST and 'target' in request.POST) or (mode == 'link')
            if is_link:
                if op == "inserted":
                    source = get_object_or_404(Tetelsor, id=request.POST.get("source"))
                    target = get_object_or_404(Tetelsor, id=request.POST.get("target"))
                    link = GanttLink.objects.create(source=source, target=target, type=request.POST.get("type", "0"))
                    return JsonResponse({"action": "inserted", "tid": link.id})
                elif op == "deleted": GanttLink.objects.filter(id=sid).delete(); return JsonResponse({"action": "deleted"})
                elif op == "updated": link = get_object_or_404(GanttLink, id=sid); link.type = request.POST.get("type", "0"); link.save(); return JsonResponse({"action": "updated"})
            else:
                if op == "updated":
                    task = get_object_or_404(Tetelsor, id=sid)
                    if "start_date" in request.POST: task.gantt_start_date = request.POST.get("start_date")[:10]
                    if "duration" in request.POST: task.gantt_duration = int(request.POST.get("duration"))
                    if "progress" in request.POST: task.progress_percentage = float(request.POST.get("progress")) * 100
                    if "felelos" in request.POST: task.felelos = request.POST.get("felelos")
                    if "owner_id" in request.POST: aid=request.POST.get("owner_id"); task.alvallalkozo = Alvallalkozo.objects.get(id=aid) if aid else None
                    task.save(); return JsonResponse({"action": "updated"})
        except: return JsonResponse({"action": "error"})
    return JsonResponse({"action": "error"})

def project_quote_html(request, pk):
    project = get_object_or_404(Project, id=pk)
    tasks = project.tetelsorok.all().order_by('sorszam')

    summary = tasks.aggregate(
        mat=Sum('anyag_osszesen'),
        ls=Sum('sajat_munkadij_osszesen'),
        la=Sum('alv_munkadij_osszesen'),
    )

    total_netto = (summary['mat'] or 0) + (summary['ls'] or 0) + (summary['la'] or 0)
    vat_amount = total_netto * (project.vat_rate / Decimal(100))
    total_brutto = total_netto + vat_amount

    company_data = get_company_context()
    context = {
        'project': project,
        'tasks': tasks,
        'total_netto': total_netto,
        'vat_amount': vat_amount,
        'total_brutto': total_brutto,
        'today': timezone.now(),
        **company_data,
    }
    return render(request, 'projects/pdf_quote.html', context)


def project_quote_pdf(request, pk):
    return project_quote_html(request, pk)


# === GLOBÁLIS GANTT ===

def global_gantt_view(request):
    alvallalkozok = Alvallalkozo.objects.all().values('id', 'nev')
    return render(request, 'projects/global_gantt.html', {'alvallalkozok': list(alvallalkozok)})


def global_gantt_data(request):
    projects = Project.objects.exclude(status='TORLES_KERELEM').order_by('start_date')
    data = []
    project_pointers = {}

    for p in projects:
        start = p.start_date or timezone.now().date()
        project_pointers[p.id] = start
        data.append(
            {'id': f"prj_{p.id}", 'text': p.name, 'tetelszam': "PROJEKT", 'start_date': start.strftime("%Y-%m-%d"),
             'type': 'project', 'open': False, 'readonly': True})

    # Itt is fontos a természetes rendezés!
    all_tasks = Tetelsor.objects.filter(project__in=projects)
    tasks = sorted(all_tasks, key=lambda t: (t.project.id, natural_sort_key(t.sorszam or "")))

    for t in tasks:
        current_pointer = project_pointers.get(t.project.id, timezone.now().date())

        calc_duration = 1
        if t.normaido and t.mennyiseg:
            try:
                hpd = float(t.project.hours_per_day or 8); calc_duration = math.ceil(
                    (float(t.mennyiseg) * float(t.normaido)) / hpd) if hpd > 0 else 1
            except:
                pass
        duration = t.gantt_duration if t.gantt_start_date and t.gantt_duration > 0 else calc_duration
        if duration < 1: duration = 1

        if t.gantt_start_date:
            start_date_obj = t.gantt_start_date
        else:
            start_date_obj = current_pointer

        start_date_obj = get_next_workday(start_date_obj)  # Hétvége javítás

        finish_date_obj = calculate_work_end_date(start_date_obj, duration)
        if finish_date_obj:
            next_start = finish_date_obj + timedelta(days=1)
            project_pointers[t.project.id] = get_next_workday(next_start)

        start_str = start_date_obj.strftime("%Y-%m-%d")
        txt = t.leiras or (t.master_item.leiras if t.master_item else "-")
        tszam = t.master_item.tetelszam if t.master_item else ""

        data.append({'id': t.id, 'text': txt, 'tetelszam': tszam, 'start_date': start_str, 'duration': duration,
                     'progress': float(t.progress_percentage) / 100, 'parent': f"prj_{t.project.id}",
                     'owner_id': t.alvallalkozo.id if t.alvallalkozo else "", 'felelos': t.felelos or ""})

    links = GanttLink.objects.filter(source__project__in=projects)
    link_data = [{'id': l.id, 'source': l.source.id, 'target': l.target.id, 'type': l.type} for l in links]
    return JsonResponse({"data": data, "links": link_data})


@csrf_exempt
def global_gantt_update(request):
    if request.method == 'POST':
        try:
            sid = request.POST.get("id")
            if str(sid).startswith("prj_"): return JsonResponse(
                {"action": "error", "msg": "A projektet nem mozgathatod!"})
            return gantt_update(request, None)
        except Exception as e:
            return JsonResponse({"action": "error", "msg": str(e)})
    return JsonResponse({"action": "error"})
@csrf_exempt
def global_gantt_update(request):
    if request.method == 'POST':
        try:
            sid = request.POST.get("id")
            # Projektet nem mozgathatunk, csak feladatot
            if str(sid).startswith("prj_"):
                return JsonResponse({"action": "error", "msg": "A projektet nem mozgathatod!"})

            # A sima gantt_update függvényt használjuk újra
            return gantt_update(request, None)
        except Exception as e:
            return JsonResponse({"action": "error", "msg": str(e)})
    return JsonResponse({"action": "error"})

def uniclass_tree_data(request):
    nodes = UniclassNode.objects.all().values('id', 'code', 'title_en', 'title_hu', 'parent_id')
    data = []
    for n in nodes:
        label = n['title_hu'] if n['title_hu'] else n['title_en']
        data.append({
            "id": str(n['id']),
            "parent": str(n['parent_id']) if n['parent_id'] else "#",
            "text": f"{n['code']} - {label}",
            "icon": "fa fa-folder" if not n['parent_id'] else "fa fa-tag",
            "a_attr": {"title": label}
        })
    return JsonResponse(data, safe=False)


@login_required
def project_chapter_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        form = ProjectChapterForm(request.POST)
        if form.is_valid():
            chapter = form.save(commit=False)
            chapter.project = project
            chapter.save()
            messages.success(request, f"'{chapter.name}' fejezet létrehozva.")
            return redirect('project-detail', pk=project.id)
    else:
        # Automatikus sorszám ajánlás (pl. ha van 10, 20, akkor 30 legyen)
        last_rank = project.chapters.order_by('-rank').first()
        initial_rank = (last_rank.rank + 10) if last_rank else 10
        form = ProjectChapterForm(initial={'rank': initial_rank})

    return render(request, 'projects/project_chapter_form.html', {'form': form, 'project': project})


@login_required
def project_chapter_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        form = ProjectChapterForm(request.POST)
        if form.is_valid():
            chapter = form.save(commit=False)
            chapter.project = project
            chapter.save()
            messages.success(request, f"'{chapter.name}' fejezet létrehozva.")
            return redirect('project-detail', pk=project.id)
    else:
        # Automatikus sorszámozás (10-esével növelve)
        last_rank = project.chapters.order_by('-rank').first()
        initial_rank = (last_rank.rank + 10) if last_rank else 10
        form = ProjectChapterForm(initial={'rank': initial_rank})

    return render(request, 'projects/project_chapter_form.html', {'form': form, 'project': project})

# === HELYŐRZŐK ===
@login_required
def resource_planning(request): return render(request, 'projects/placeholder.html',
                                              {'title': 'Éves Erőforrás Ütemezés'})


@login_required
def hr_dashboard(request):
    # 1. Hónap kiválasztása (Alapértelmezett: aktuális hónap)
    today = timezone.now().date()
    selected_month_str = request.GET.get('month')

    if selected_month_str:
        try:
            # Biztonságos dátum konverzió
            selected_date = datetime.strptime(selected_month_str, "%Y-%m").date()
        except ValueError:
            selected_date = today.replace(day=1)
    else:
        selected_date = today.replace(day=1)

    # 2. Dolgozók lekérése
    employees = Employee.objects.filter(status='ACTIVE')
    payroll_data = []

    for emp in employees:
        # a) Ledolgozott órák ebben a hónapban
        attendance_qs = Attendance.objects.filter(
            employee=emp,
            date__year=selected_date.year,
            date__month=selected_date.month
        )
        total_hours = attendance_qs.aggregate(Sum('hours_worked'))['hours_worked__sum'] or 0

        # b) Munkabér számítása (Napi bér / 8 * ledolgozott óra)
        # Ha nincs napi bér beállítva, 0-val számolunk
        daily_cost = emp.daily_cost or 0
        hourly_rate = Decimal(daily_cost) / Decimal(8)
        base_salary = Decimal(total_hours) * hourly_rate

        # c) Pénzügyi tételek (Előleg, Prémium)
        items = PayrollItem.objects.filter(
            employee=emp,
            date__year=selected_date.year,
            date__month=selected_date.month
        )

        advances = items.filter(type='ADVANCE').aggregate(Sum('amount'))['amount__sum'] or 0
        prems = items.filter(type='PREMIUM').aggregate(Sum('amount'))['amount__sum'] or 0
        deductions = items.filter(type='DEDUCTION').aggregate(Sum('amount'))['amount__sum'] or 0

        # d) Végösszeg (Alapbér + Prémium - Előleg - Levonás)
        final_pay = base_salary + prems - advances - deductions

        payroll_data.append({
            'employee': emp,
            'hours': total_hours,
            'base_salary': base_salary,
            'advances': advances,
            'premiums': prems,
            'deductions': deductions,
            'final_pay': final_pay
        })

    context = {
        'payroll_data': payroll_data,
        'selected_date': selected_date,
        'today': today
    }
    return render(request, 'projects/hr_dashboard.html', context)

@login_required
def global_inventory(request): all_items = ProjectInventory.objects.all().order_by('name'); return render(request,
                                                                                                          'projects/global_inventory.html',
                                                                                                          {
                                                                                                              'inventory': all_items})


@login_required
def finance_dashboard(request): total = Expense.objects.aggregate(Sum('amount_netto'))[
                                            'amount_netto__sum'] or 0; return render(request,
                                                                                     'projects/placeholder.html',
                                                                                     {'title': 'Pénzügyi Kimutatások',
                                                                                      'total': total})


@login_required
def asset_list(request): return render(request, 'projects/placeholder.html', {'title': '🚜 Géppark'})


@login_required
def project_map_view(request): return render(request, 'projects/placeholder.html', {'title': '🗺️ Térkép'})


@login_required
def crm_dashboard(request): return render(request, 'projects/placeholder.html', {'title': '🤝 CRM'})