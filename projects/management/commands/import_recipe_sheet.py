# projects/management/commands/import_recipe_sheet.py

from django.core.management.base import BaseCommand
from projects.models import (
    MasterItem, UniclassNode, Material, Operation, Machine,
    ItemComponent, LaborComponent, MachineComponent
)
import openpyxl
import os
from decimal import Decimal


class Command(BaseCommand):
    help = 'Egyedi Tétel Receptúra (Norma lap) importálása Excelből'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Az Excel fájl elérési útja')

    def handle(self, *args, **options):
        file_path = options['file_path']

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'HIBA: Nincs ilyen fájl: {file_path}'))
            return

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # 1. FEJLÉC ADATOK
        # A feltöltött fájl alapján:
        # B1: Uniclass kód (Ef_25_10...)
        # B2: Tételszám
        # B3: Megnevezés

        uniclass_code_raw = sheet['B1'].value
        tetelszam = sheet['B2'].value
        leiras = sheet['B3'].value

        if not tetelszam or not leiras:
            self.stdout.write(self.style.ERROR("HIBA: Hiányzó fejléc adatok (B2/B3)!"))
            return

        # Uniclass keresése (Levágjuk a zárójelet)
        # Pl. "Ef_25_10 (Walls)" -> "Ef_25_10"
        uniclass_code = str(uniclass_code_raw).split('(')[0].strip()
        uniclass_node = UniclassNode.objects.filter(code=uniclass_code).first()

        if not uniclass_node:
            self.stdout.write(self.style.WARNING(f"Figyelem: Uniclass kód nem található: {uniclass_code}"))

        # MASTER ITEM MENTÉSE
        master_item, created = MasterItem.objects.update_or_create(
            tetelszam=tetelszam,
            defaults={
                'leiras': leiras,
                'egyseg': 'm2',  # Alapértelmezés, vagy kiolvasható a B13 környékéről
                'uniclass_item': uniclass_node
            }
        )

        self.stdout.write(f"--- Tétel: {tetelszam} ({'Létrehozva' if created else 'Frissítve'}) ---")

        # Előző komponensek törlése (tiszta lappal indítunk)
        ItemComponent.objects.filter(master_item=master_item).delete()
        LaborComponent.objects.filter(master_item=master_item).delete()
        MachineComponent.objects.filter(master_item=master_item).delete()

        # 2. SOROK FELDOLGOZÁSA
        # A mintafájlban a 9. sor körül kezdődik a táblázat
        # Keressük meg a fejlécet: "Típus"

        start_row = 1
        for row in sheet.iter_rows(min_row=1, max_row=20):
            if row[0].value == "Típus":
                start_row = row[0].row + 1
                break

        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            tipus = row[0]
            nev = row[1]
            mennyiseg = row[2]
            egyseg = row[3]
            ar = row[4]

            # Ha üres a sor vagy vége a táblázatnak ("Tétel összesen")
            if not tipus or str(tipus).startswith("Tétel összesen"):
                continue

            # Számok tisztítása
            try:
                mennyiseg = Decimal(str(mennyiseg).replace(',', '.')) if mennyiseg else 0
                ar = Decimal(str(ar).replace(',', '.')) if ar else 0
            except:
                continue

            tipus_str = str(tipus).lower().strip()

            # A) MUNKA
            if 'munka' in tipus_str:
                op, _ = Operation.objects.get_or_create(name=nev, defaults={'hourly_rate': ar})
                LaborComponent.objects.create(master_item=master_item, operation=op, time_required=mennyiseg)
                self.stdout.write(f"  + Munka: {nev}")

            # B) ANYAG
            elif 'anyag' in tipus_str:
                mat, _ = Material.objects.get_or_create(name=nev, defaults={'price': ar, 'unit': egyseg})
                ItemComponent.objects.create(master_item=master_item, material=mat, amount=mennyiseg)
                self.stdout.write(f"  + Anyag: {nev}")

            # C) GÉP
            elif 'gép' in tipus_str or 'gep' in tipus_str:
                mac, _ = Machine.objects.get_or_create(name=nev, defaults={'price': ar, 'unit': egyseg})
                MachineComponent.objects.create(master_item=master_item, machine=mac, amount=mennyiseg)
                self.stdout.write(f"  + Gép: {nev}")

        # Árak újraszámolása
        master_item.calculate_totals()
        self.stdout.write(self.style.SUCCESS(f"SIKER! Tétel importálva: {master_item.total_price} Ft"))