# projects/management/commands/import_bulk_master.py

from django.core.management.base import BaseCommand
from django.db import transaction
from projects.models import (
    MasterItem, UniclassNode, Material, Operation, Machine,
    ItemComponent, LaborComponent, MachineComponent
)
import openpyxl
import os
from decimal import Decimal


class Command(BaseCommand):
    help = 'Csoportos Törzstétel és Receptúra Importálás (4 munkalapos Excel)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Az Excel fájl elérési útja')

    def handle(self, *args, **options):
        file_path = options['file_path']

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'HIBA: Nincs ilyen fájl: {file_path}'))
            return

        self.stdout.write("--- Excel megnyitása... ---")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Hiba a fájl megnyitásakor: {e}'))
            return

        # Ellenőrizzük a munkalapokat
        required_sheets = ['Tetelek', 'Anyagok', 'Munka', 'Gepek']
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                self.stdout.write(self.style.ERROR(f"HIBA: Hiányzik a '{sheet}' munkalap!"))
                return

        stats = {'items': 0, 'materials': 0, 'labor': 0, 'machines': 0}
        master_cache = {}  # Memóriában tároljuk a tételeket a gyors eléréshez

        try:
            with transaction.atomic():  # Ha hiba van, semmi nem mentődik el (biztonságos)

                # === 1. TÉTELEK (MASTER ITEMS) ===
                self.stdout.write("1. Tételek betöltése...")
                ws = wb['Tetelek']
                # Feltételezett oszlopok: A=Tételszám, B=Név, C=Egység, D=Uniclass

                for row in ws.iter_rows(min_row=2, values_only=True):
                    code = str(row[0]).strip() if row[0] else None
                    if not code: continue

                    name = row[1]
                    unit = row[2] or "db"
                    uniclass_code = row[3]

                    # Uniclass keresése
                    uni_node = None
                    if uniclass_code:
                        uni_node = UniclassNode.objects.filter(code=str(uniclass_code).strip()).first()

                    # Tétel létrehozása vagy frissítése
                    item, created = MasterItem.objects.update_or_create(
                        tetelszam=code,
                        defaults={
                            'leiras': name,
                            'egyseg': unit,
                            'uniclass_item': uni_node
                        }
                    )

                    # Töröljük a régi receptet, hogy tisztán építsük újra
                    ItemComponent.objects.filter(master_item=item).delete()
                    LaborComponent.objects.filter(master_item=item).delete()
                    MachineComponent.objects.filter(master_item=item).delete()

                    master_cache[code] = item
                    stats['items'] += 1

                # === 2. ANYAGOK ===
                self.stdout.write("2. Anyagok csatolása...")
                ws = wb['Anyagok']
                # Oszlopok: A=Tételszám, B=AnyagNév, C=Mennyiség, D=Egység, E=Ár

                for row in ws.iter_rows(min_row=2, values_only=True):
                    t_code = str(row[0]).strip() if row[0] else None
                    if not t_code or t_code not in master_cache: continue

                    m_name = row[1]
                    amount = self.clean_decimal(row[2])
                    unit = row[3] or "db"
                    price = self.clean_decimal(row[4])

                    # Anyag létrehozása, ha nem létezik
                    mat, _ = Material.objects.get_or_create(
                        name=m_name,
                        defaults={'unit': unit, 'price': price}
                    )

                    ItemComponent.objects.create(
                        master_item=master_cache[t_code],
                        material=mat,
                        amount=amount
                    )
                    stats['materials'] += 1

                # === 3. MUNKA ===
                self.stdout.write("3. Munkaerő csatolása...")
                ws = wb['Munka']
                # Oszlopok: A=Tételszám, B=MűveletNév, C=NormaIdő, D=Rezsióradíj

                for row in ws.iter_rows(min_row=2, values_only=True):
                    t_code = str(row[0]).strip() if row[0] else None
                    if not t_code or t_code not in master_cache: continue

                    op_name = row[1]
                    time_req = self.clean_decimal(row[2])
                    rate = self.clean_decimal(row[3])

                    op, _ = Operation.objects.get_or_create(
                        name=op_name,
                        defaults={'hourly_rate': rate}
                    )

                    LaborComponent.objects.create(
                        master_item=master_cache[t_code],
                        operation=op,
                        time_required=time_req
                    )
                    stats['labor'] += 1

                # === 4. GÉPEK ===
                self.stdout.write("4. Gépek csatolása...")
                ws = wb['Gepek']
                # Oszlopok: A=Tételszám, B=GépNév, C=Használat, D=Költség

                for row in ws.iter_rows(min_row=2, values_only=True):
                    t_code = str(row[0]).strip() if row[0] else None
                    if not t_code or t_code not in master_cache: continue

                    mac_name = row[1]
                    amount = self.clean_decimal(row[2])
                    price = self.clean_decimal(row[3])

                    mac, _ = Machine.objects.get_or_create(
                        name=mac_name,
                        defaults={'price': price}
                    )

                    MachineComponent.objects.create(
                        master_item=master_cache[t_code],
                        machine=mac,
                        amount=amount
                    )
                    stats['machines'] += 1

                # === 5. ÚJRASZÁMOLÁS ===
                self.stdout.write("Árak frissítése...")
                for item in master_cache.values():
                    item.calculate_totals()

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Kritikus hiba: {e}"))
            return

        self.stdout.write(self.style.SUCCESS(
            f"SIKERES IMPORT!\n"
            f"--------------------------------\n"
            f"Tételek: {stats['items']}\n"
            f"Anyag sorok: {stats['materials']}\n"
            f"Munka sorok: {stats['labor']}\n"
            f"Gép sorok: {stats['machines']}"
        ))

    def clean_decimal(self, value):
        try:
            if value is None: return Decimal(0)
            return Decimal(str(value).replace(',', '.'))
        except:
            return Decimal(0)