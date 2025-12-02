import openpyxl
import random

# Fájlnév
FILENAME = "bulk_data_example.xlsx"

# --- ADATGENERÁTOR LISTÁK ---
CATEGORIES = [
    ("FAL", "Falazás", "m2", "Ef_25_10"),
    ("BETON", "Betonozás", "m3", "Ef_20_10"),
    ("VAK", "Vakolás", "m2", "Ef_25_10"),
    ("BURK", "Hidegburkolás", "m2", "Ss_25_10"),
    ("FEST", "Festés", "m2", "Ss_25_12"),
]

MATERIALS = [
    ("Porotherm 30 N+F", "db", 850), ("Zsákos Esztrich", "kg", 120),
    ("Csemperagasztó Flex", "kg", 450), ("Diszperziós festék", "l", 1200),
    ("Betonacél 8mm", "kg", 400), ("Zsaludeszka", "m3", 150000),
    ("Homok", "m3", 8000), ("Cement", "kg", 60), ("Fugaanyag", "kg", 1500)
]

OPERATIONS = [
    ("Kőműves munka", 6500), ("Segédmunka", 4500),
    ("Burkoló munka", 7500), ("Festő munka", 6000),
    ("Ács munka", 7000), ("Betonozás", 5500)
]

MACHINES = [
    ("Betonkeverő", 3000), ("Daru", 15000),
    ("Vakológép", 5000), ("Hilti Vésőgép", 2500),
    ("Állványzat", 1200)
]

# --- MUNKAÜZET LÉTREHOZÁSA ---
wb = openpyxl.Workbook()

# 1. Munkalap: Tetelek
ws_items = wb.active
ws_items.title = "Tetelek"
ws_items.append(["Tételszám", "Megnevezés", "Egység", "Uniclass Kód"])

# 2. Munkalap: Anyagok
ws_mat = wb.create_sheet("Anyagok")
ws_mat.append(["Tételszám", "Anyag neve", "Mennyiség", "Egység", "Egységár"])

# 3. Munkalap: Munka
ws_lab = wb.create_sheet("Munka")
ws_lab.append(["Tételszám", "Művelet neve", "Norma idő", "Rezsióradíj"])

# 4. Munkalap: Gepek
ws_mac = wb.create_sheet("Gepek")
ws_mac.append(["Tételszám", "Gép neve", "Használat", "Költség"])

print(f"Generálás indítása: 100 tétel...")

for i in range(1, 101):
    # Tétel generálás
    cat_code, cat_name, unit, uni_code = random.choice(CATEGORIES)
    item_code = f"{cat_code}-{i:03d}"
    item_name = f"{cat_name} {i}. típus (Generált)"

    ws_items.append([item_code, item_name, unit, uni_code])

    # Anyagok (1-3 db per tétel)
    for _ in range(random.randint(1, 3)):
        mat_name, mat_unit, mat_price = random.choice(MATERIALS)
        amount = round(random.uniform(1.0, 20.0), 2)
        # Tételszám összekapcsolása
        ws_mat.append([item_code, mat_name, amount, mat_unit, mat_price])

    # Munka (1-2 db per tétel)
    for _ in range(random.randint(1, 2)):
        op_name, op_rate = random.choice(OPERATIONS)
        time_req = round(random.uniform(0.5, 5.0), 2)
        # Tételszám összekapcsolása
        ws_lab.append([item_code, op_name, time_req, op_rate])

    # Gép (0-1 db per tétel)
    if random.random() > 0.5:  # 50% esély
        mac_name, mac_price = random.choice(MACHINES)
        usage = round(random.uniform(0.1, 2.0), 2)
        # Tételszám összekapcsolása
        ws_mac.append([item_code, mac_name, usage, mac_price])

# Mentés
wb.save(FILENAME)
print(f"KÉSZ! A fájl létrejött: {FILENAME}")